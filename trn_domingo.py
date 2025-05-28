import oracledb
import os
import sys
from datetime import datetime, timedelta
import configparser
import openpyxl
from collections import Counter
from logs_escrita import (log_info, log_warning, log_debug,  log_error) # noqa

# Defini versão
usuario = os.getlogin()

# Variaveis
hora_atual = datetime.now()
nm_log_data = datetime.strftime(datetime.now(), '%d_%m_%Y')
hoje = datetime.now()
dia = hoje.weekday()  # 0=segunda, 6=domingo
ontem = hoje - timedelta(days=1)
ontem = ontem.strftime('%d/%m/%Y')


lista_trn = []
lista_opc = []
lista_s_trn = []
expurgo_final = []


# Caminho do Parser
def get_app_and_settings_full_path():
    if getattr(sys, 'frozen', False):
        BASE_PATH = os.path.dirname(sys.executable)
    else:
        BASE_PATH = os.path.dirname(os.path.abspath(__file__))
    return BASE_PATH, os.path.join(BASE_PATH, "Config.ini")


CAM_LOGS_LOGS, CAM_CONFIG_PARSER = get_app_and_settings_full_path()
CAMINHO_LOGS = f'{CAM_LOGS_LOGS}\\DADOS'

# Criar objeto do configparser
config = configparser.ConfigParser()
with open(CAM_CONFIG_PARSER, "r", encoding="utf-8") as file:
    config.read_file(file)

# Ler o arquivo ini
ambiente = config["ambiente"]["ambiente"]

# Acessar os valores das seções e chaves
LOG_ESCRITA = config[ambiente]["log"]
ARQ_OPC = config[ambiente]["opc"]
ARQ_EXP = config[ambiente]["expurgo"]
USERNAME = config[ambiente]["user"]
SENHA = config[ambiente]["password"]
CON_DSN = config[ambiente]["dsn"]
CON_PORT = config[ambiente]["port"]
CON_SERVICE = config[ambiente]["service"]
DIAEXECU = config[ambiente]["dia"]


def opc_dados():
    # Carrega a Planilha
    workbook = openpyxl.load_workbook(ARQ_OPC)
    sheet = workbook["Page 1"]
    sheet = workbook.active

    # Acessando os dados
    for row in sheet.iter_rows(min_row=0, max_row=sheet.max_row, min_col=1, max_col=16): # type: ignore # noqa
        codigo = row[0].value
        servico = row[15].value

        if servico == "Abastece":
            lista_remover = [" - V4 NUC", " - SPP", " - 3"]
            for item in lista_remover:
                codigo = codigo.replace(item, "")  # type: ignore
            lista_opc.append(codigo)  # type: ignore

    workbook.close()

    if LOG_ESCRITA == 'Sim':
        msg_log = f'O OPC foi carregado com sucesso, {len(lista_opc)} postos.'
        log_info(msg_log)

    # print(nova_lista)

    # Salva a planilha
    workbook = openpyxl.Workbook()  # type: ignore
    sheet = workbook.active
    workbook.title = "Page 1"  # type: ignore

    # Escreve os postos em outra lista
    for inx, valor in enumerate(lista_opc, start=1):
        sheet.cell(row=inx, column=1, value=valor)  # type: ignore

    workbook.save(f'{CAMINHO_LOGS}\\OPC_{nm_log_data}.xlsx')
    print('Concluído!')


def executa():
    # Ativa o modo de compatibilidade com cliente leve (sem Instant Client)
    if not oracledb.is_thin_mode():
        oracledb.init_oracle_client(lib_dir="instantclient_23_8")

    conn = oracledb.connect(
        user=USERNAME,
        password=SENHA,
        dsn=f'{CON_DSN}:{CON_PORT}/{CON_SERVICE}'
    )

    resultado = conn.cursor()

    sql = """
    SELECT TB0008_CD_CONVENIADO AS Postos,
        TO_CHAR(tb0153_dt_transacao, 'YYYY-MM-DD') AS domingo,
        COUNT(*) AS quantidade_registros
    FROM tb0153_transacaoconveniado
    WHERE tb0153_dt_transacao >= TRUNC(SYSDATE, 'IW') - 21
    AND TO_CHAR(
    tb0153_dt_transacao, 'DY', 'NLS_DATE_LANGUAGE=PORTUGUESE') = 'DOM'
    AND tb0138_cd_produto = '1'
    GROUP BY TB0008_CD_CONVENIADO, TO_CHAR(tb0153_dt_transacao, 'YYYY-MM-DD')
    ORDER BY domingo DESC
    """

    resultado.execute(sql)

    for cod, data_tr, qtd in resultado:
        # print(f"Posto: {cod}, Data: {data_tr}, Quantidade: {qtd}")
        lista_trn.append(cod)

    resultado.close()
    conn.close()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    workbook.title = 'trn'  # type: ignore

    for inx, valor in enumerate(lista_trn, start=1):
        sheet.cell(row=inx, column=1, value=valor)  # type: ignore

    workbook.save(f'{CAMINHO_LOGS}\\TRN_{nm_log_data}.xlsx')

    if LOG_ESCRITA == 'Sim':
        msg_log = f'As transações foram carregadas com sucesso, {len(lista_trn)} postos.' # noqa
        log_info(msg_log)


def gera_expurgo():
    global lista_opc, lista_trn, lista_s_trn, expurgo_final

    lista_opc = [str(c) for c in lista_opc]
    lista_trn = [str(c) for c in lista_trn]

    contagem_domingo = Counter(lista_trn)

    for codigo_uni in lista_opc:
        quantidade = contagem_domingo.get(codigo_uni, 0)
        if quantidade == 0:
            lista_s_trn.append(int(codigo_uni))

    # Adiciona info
    for i in lista_s_trn:
        expurgo_final.append((i, "Não", ontem, "Sim"))
        # print(expurgo_final)

    print(ARQ_EXP)
    workbook = openpyxl.load_workbook(ARQ_EXP)
    sheet = workbook["Abono"]

    l_inicial = sheet.max_row + 1  # type: ignore
    for i, linha in enumerate(expurgo_final, start=l_inicial):
        for j, valor in enumerate(linha, start=1):
            sheet.cell(row=i, column=j, value=valor)  # type: ignore

    workbook.save(ARQ_EXP)

    if LOG_ESCRITA == 'Sim':
        msg_log = f'Os abonos foram adicionados com sucesso a planilha, {ARQ_EXP}.' # noqa
        log_info(msg_log)


# Valida o dia da semana
def dia_da_semana():
    if dia == int(DIAEXECU):  # Se for domingo
        if LOG_ESCRITA == 'Sim':
            msg_log = f'A rotina de criação de expurgos iniciou' # noqa
            log_info(msg_log)

        executa()
        opc_dados()
        gera_expurgo()
        print(f"Hoje é segunda, {dia}")

        if LOG_ESCRITA == 'Sim':
            msg_log = f'A rotina de criação de expurgos concluíu!' # noqa
            log_info(msg_log)

        return True
    else:
        if LOG_ESCRITA == 'Sim':
            msg_log = f'A rotina não executará hoje, programado para Segunda!' # noqa
            log_info(msg_log)

        print(f"Hoje não é segunda, {dia}")
        return False


if __name__ == "__main__":
    dia_da_semana()
